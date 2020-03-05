# -*- coding:utf-8 _*-
#!/usr/bin/python3

#   author : YOYO
#   time :  2020/3/5 20:10
#   email : youyou.xu@enesource.com
#   project_name :  test_read_mysql 
#   file_name :  do_excel.py
#   function： 从Excel里读取测试用例，保存成列表中的对象

import common.contants
from openpyxl import load_workbook   # 操作Excel必须要引用的库



class Case:
    def __init__(self):
        self.id = None
        self.title = None
        self.api = None
        self.data = None
        self.expected = None


class DoExcel:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def open_sheet(self):
        excel = load_workbook(self.filename) # 打开Excel
        sheet = excel[self.sheetname]  # 打开sheet
        return excel, sheet

    def read_data(self):
        sheet = self.open_sheet()[1]
        cases = []
        for i in range(2, sheet.max_row +1 ):
            dictdata = {} # 请求参数用字典存储
            case = Case()
            case.id = sheet.cell(i, 1).value
            case.title = sheet.cell(i, 2).value
            case.api = sheet.cell(i, 3).value
            case.expected = sheet.cell(i, 5).value

            dictdata['data'] = sheet.cell(i, 4).value
            case.data = dictdata
            cases.append(case)
        return cases

    def write_data(self, row, column, value):
        excel, sheet = self.open_sheet()
        sheet.cell(row, column).value = value
        excel.save(self.filename)


if __name__ =="__main__":
    case_data = common.contants.case_data
    excel = DoExcel(case_data, "Sheet1").read_data()
    for i in excel:
        print( i.expected)




